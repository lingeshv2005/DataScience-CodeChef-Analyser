import requests
from bs4 import BeautifulSoup
import urllib.parse
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
import logging
import openpyxl
from pathlib import Path
import requests.exceptions
import random

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def save_to_excel(users, badges_all, ratings_all, ranks_all, submissions_all, file_path):
    try:
        workbook = openpyxl.Workbook()
        
        # Users sheet
        ws_users = workbook.active
        ws_users.title = "Users"
        ws_users.append(["Username"])
        for username in users:
            ws_users.append([username])
        
        # Badges sheet
        ws_badges = workbook.create_sheet("Badges")
        ws_badges.append(["Username", "Title", "Description", "Image URL"])
        for b in badges_all:
            ws_badges.append(b)
        
        # Ratings sheet
        ws_ratings = workbook.create_sheet("Ratings")
        ws_ratings.append(["Username", "Rating", "Stars", "Highest"])
        for r in ratings_all:
            ws_ratings.append(r)
        
        # Ranks sheet
        ws_ranks = workbook.create_sheet("Ranks")
        ws_ranks.append(["Username", "Label", "Rank"])
        for rk in ranks_all:
            ws_ranks.append(rk)
        
        # Submissions sheet
        ws_subs = workbook.create_sheet("Submissions")
        ws_subs.append(["Username", "Time", "Problem", "Result", "Language", "Solution Link"])
        for s in submissions_all:
            ws_subs.append(s)
        
        # Save the file
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(file_path)
        logging.info(f"Saved data for {len(users)} users to {file_path}")
    except Exception as e:
        logging.error(f"Failed to save to Excel file {file_path}: {e}")

def get_usernames_from_institution(institution, max_pages_limit=20, max_retries=3):
    usernames = set()
    institution_encoded = urllib.parse.quote(institution)
    page = 1

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--log-level=3")

    driver = webdriver.Chrome(options=chrome_options)

    try:
        while page <= max_pages_limit:
            url = (
                f"https://www.codechef.com/ratings/all"
                f"?filterBy=Institution%3D{institution_encoded}"
                f"&itemsPerPage=50&order=asc&page={page}&sortBy=global_rank"
            )
            logging.info(f"Fetching page {page}: {url}")

            for attempt in range(max_retries):
                try:
                    driver.get(url)
                    wait = WebDriverWait(driver, 30)
                    wait.until_not(EC.presence_of_element_located((By.CLASS_NAME, "loadingIcon")))
                    time.sleep(2)
                    break
                except Exception as e:
                    logging.warning(f"Attempt {attempt + 1}/{max_retries} failed for page {page}: {e}")
                    if attempt + 1 == max_retries:
                        logging.error(f"Max retries reached for page {page}, skipping")
                        return sorted(list(usernames))
                    time.sleep(2 ** attempt + random.uniform(5.0, 8.0))

            soup = BeautifulSoup(driver.page_source, "html.parser")
            page_text = soup.get_text().lower()
            if "no results" in page_text or "no users" in page_text or "0 results" in page_text:
                logging.info(f"No results found on page {page}")
                break

            table = soup.find("table", class_=re.compile(r"MuiTable-root.*MUIDataTable-tableRoot")) or soup.find("table")
            if not table:
                logging.error(f"No table found on page {page}. Page title: {soup.title.get_text() if soup.title else 'No title'}")
                break

            rows = table.find_all("tr")
            if len(rows) <= 1:
                logging.warning(f"No data rows found on page {page}")
                break

            page_usernames = set()
            for row in rows[1:]:
                first_td = row.find("td", {"data-colindex": "0"})
                if first_td:
                    link = first_td.find("a", href=re.compile(r"/users/"))
                    if link:
                        username_span = link.find("span", class_="m-username--link")
                        username = username_span.get_text(strip=True) if username_span else link.get("title", "")
                        if username:
                            page_usernames.add(username)

            new_users = len(page_usernames - usernames)
            usernames.update(page_usernames)
            logging.info(f"Extracted {len(page_usernames)} users from page {page} ({new_users} new). Total so far: {len(usernames)}")

            if new_users == 0:
                logging.info(f"No new users on page {page}, stopping")
                break
            if len(page_usernames) < 50 and page > 1:
                logging.info(f"Partial page {page} ({len(page_usernames)} users < 50), likely last page")
                break

            next_button = soup.find("button", {"aria-label": re.compile(r"Go to next page", re.I)}) or \
                          soup.find("button", string=re.compile(r"Next", re.I)) or \
                          soup.find("button", class_=re.compile(r"MuiPaginationItem.*next"))
            if not next_button or ("disabled" in next_button.get("class", []) or next_button.get("disabled") or "false" in next_button.get("aria-disabled", "")):
                logging.info("Next button disabled or not found - no more pages")
                break

            page += 1
            time.sleep(random.uniform(5.0, 8.0))

    except Exception as e:
        logging.error(f"Error during extraction: {e}")
    finally:
        driver.quit()

    return sorted(list(usernames))

def scrape_user_profile(handle, session, headers, max_retries=3):
    badges_data = []
    ratings_data = []
    ranks_data = []
    submissions_data = []

    # Scrape profile page
    url = f"https://www.codechef.com/users/{handle}"
    for attempt in range(max_retries):
        try:
            response = session.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                break
            logging.warning(f"Attempt {attempt + 1}/{max_retries} failed for profile {handle}: Status {response.status_code}")
            if attempt + 1 == max_retries:
                logging.error(f"Max retries reached for profile {handle}, skipping")
                return badges_data, ratings_data, ranks_data, submissions_data
            time.sleep(2 ** attempt + random.uniform(5.0, 8.0))
        except requests.exceptions.RequestException as e:
            logging.warning(f"Attempt {attempt + 1}/{max_retries} failed for profile {handle}: {e}")
            if attempt + 1 == max_retries:
                logging.error(f"Max retries reached for profile {handle}, skipping")
                return badges_data, ratings_data, ranks_data, submissions_data
            time.sleep(2 ** attempt + random.uniform(5.0, 8.0))

    soup = BeautifulSoup(response.text, 'html.parser')

    # Badges
    badge_section = soup.find_all('div', class_='widget badges')
    for section in badge_section:
        badges = section.find_all('div', class_='badge')
        for badge in badges:
            title = badge.find('p', class_='badge__title').text.strip() if badge.find('p', class_='badge__title') else "N/A"
            description = badge.find('p', class_='badge__description').text.strip() if badge.find('p', class_='badge__description') else "N/A"
            image_url = badge.find('img')['src'] if badge.find('img') and 'src' in badge.find('img').attrs else "N/A"
            badges_data.append([handle, title, description, image_url])

    # Rating & Stars & Highest
    rating_section = soup.find('div', class_='rating-header text-center')
    rating, star, highest = "N/A", "N/A", "N/A"
    if rating_section:
        rating_elem = rating_section.find('div', class_='rating-number')
        star_elem = rating_section.find('span')
        highest_elem = rating_section.find('small')
        rating = rating_elem.text.strip() if rating_elem else "N/A"
        star = star_elem.text.strip() if star_elem else "N/A"
        highest = highest_elem.text.strip() if highest_elem else "N/A"
    ratings_data = [handle, rating, star, highest]

    # Ranks
    ranks_section = soup.find('div', class_='rating-ranks')
    if ranks_section:
        list_items = ranks_section.find_all('li')
        for item in list_items:
            rank = item.find('strong').text.strip() if item.find('strong') else "N/A"
            label = item.get_text(strip=True).replace(rank, "").strip()
            ranks_data.append([handle, label, rank])

    # Recent Submissions (via API)
    api_url = f"https://www.codechef.com/recent/user?user_handle={handle}&page=0"
    for attempt in range(max_retries):
        try:
            resp = session.get(api_url, headers=headers, timeout=10)
            if resp.status_code == 200:
                break
            logging.warning(f"Attempt {attempt + 1}/{max_retries} failed for submissions {handle}: Status {resp.status_code}")
            if attempt + 1 == max_retries:
                logging.error(f"Max retries reached for submissions {handle}, skipping")
                return badges_data, ratings_data, ranks_data, submissions_data
            time.sleep(2 ** attempt + random.uniform(5.0, 8.0))
        except requests.exceptions.RequestException as e:
            logging.warning(f"Attempt {attempt + 1}/{max_retries} failed for submissions {handle}: {e}")
            if attempt + 1 == max_retries:
                logging.error(f"Max retries reached for submissions {handle}, skipping")
                return badges_data, ratings_data, ranks_data, submissions_data

    try:
        data = resp.json()
        html_content = data.get("content", "")
        soup_sub = BeautifulSoup(html_content, "html.parser")
        rows = soup_sub.find_all("tr")
        for row in rows:
            cols = row.find_all("td")
            if len(cols) == 5:
                time_val = cols[0].text.strip()
                problem = cols[1].text.strip()
                result = cols[2].text.strip()
                language = cols[3].text.strip()
                link_tag = cols[4].find("a")
                solution_link = f"https://www.codechef.com{link_tag['href']}" if link_tag and link_tag.has_attr("href") else "N/A"
                submissions_data.append([handle, time_val, problem, result, language, solution_link])
    except ValueError as e:
        logging.warning(f"Failed to parse JSON for submissions of {handle}: {e}")

    return badges_data, ratings_data, ranks_data, submissions_data

# ------------------ MAIN ------------------ #
if __name__ == "__main__":
    institution = "Sri Eshwar College of Engineering, Kinathukadavu"
    output_file = r"C:\AllOther\Python\CodeChef\codechefprofiles.xlsx"
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36'}
    
    # Optional: Proxy support (uncomment and configure if needed)
    # proxies = {
    #     'http': 'http://your_proxy:port',
    #     'https': 'https://your_proxy:port'
    # }
    
    session = requests.Session()
    # session.proxies = proxies  # Uncomment if using proxies

    users = get_usernames_from_institution(institution)
    
    if not users:
        logging.warning(f"No users found for {institution}, skipping Excel file creation")
        print(f"No users found for {institution}")
    else:
        logging.info(f"Found {len(users)} users from {institution}")
        print(f"✅ Found {len(users)} users from {institution}")
        print(users)
        
        badges_all = []
        ratings_all = []
        ranks_all = []
        submissions_all = []
        
        for handle in users:
            logging.info(f"Scraping profile for {handle}")
            badges, ratings, ranks, subs = scrape_user_profile(handle, session, headers)
            badges_all.extend(badges)
            if ratings and ratings[1] != "N/A":
                ratings_all.append(ratings)
            ranks_all.extend(ranks)
            submissions_all.extend(subs)
            time.sleep(random.uniform(5.0, 8.0))  # Increased random delay to avoid rate-limiting
            
        save_to_excel(users, badges_all, ratings_all, ranks_all, submissions_all, output_file)